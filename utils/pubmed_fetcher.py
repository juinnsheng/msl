"""
pubmed_fetcher.py
Fetches and parses PubMed records via NCBI E-utilities.
All network calls isolated here — no API keys leak to templates.
"""

import time
import io
import requests
import xml.etree.ElementTree as ET
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os


# ── Module-level config (set from app.py) ──────────────────────
NCBI_API_KEY: str = ""
_DELAY: float = 0.35

_ESEARCH = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
_EFETCH  = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"

def get_ncbi_api_key():
    ncbi_key = os.environ.get("NCBI_API_KEY")
    if not ncbi_key:
        raise RuntimeError("NCBI_API_KEY is not set")
    return ncbi_key

def _base_params() -> dict:
    p = {
        "db": "pubmed",
        "retmode": "xml",
        "tool": "msl_app",
        "email": "msl@example.com",
    }

    try:
        p["api_key"] = get_ncbi_api_key()
    except RuntimeError:
        pass  # allow running without key (just slower rate limit)

    return p


def esearch(query: str, max_results: int = 1000, sort: str = "relevance") -> dict:
    params = {**_base_params(), "term": query, "retmax": max_results,
              "sort": sort, "usehistory": "y"}
    r = requests.get(_ESEARCH, params=params, timeout=30)
    r.raise_for_status()
    root = ET.fromstring(r.text)
    pmids = [e.text for e in root.findall(".//Id")]
    total = int(root.findtext("Count") or 0)
    trans = root.findtext("QueryTranslation") or query
    webenv = root.findtext("WebEnv") or ""
    qkey   = root.findtext("QueryKey") or ""
    return {"pmids": pmids, "total_count": total, "query_translation": trans,
            "webenv": webenv, "query_key": qkey}


def efetch_full(pmids: list, batch: int = 200) -> list:
    records = []
    for i in range(0, len(pmids), batch):
        chunk = pmids[i:i + batch]
        params = {**_base_params(), "id": ",".join(chunk), "rettype": "xml"}
        r = requests.get(_EFETCH, params=params, timeout=60)
        r.raise_for_status()
        records.extend(_parse_xml(r.text))
        time.sleep(_DELAY)
    return records


def _parse_xml(xml_text: str) -> list:
    root = ET.fromstring(xml_text)
    out = []
    for article in root.findall("PubmedArticle"):
        med = article.find("MedlineCitation")
        art = med.find("Article") if med is not None else None
        if art is None:
            continue
        rec = {}
        rec["pmid"]  = med.findtext("PMID") or ""
        rec["title"] = "".join(art.find("ArticleTitle").itertext()) if art.find("ArticleTitle") is not None else ""

        abstract_el = art.find("Abstract")
        abstract_sections = {}
        abstract_full = ""
        if abstract_el is not None:
            for txt in abstract_el.findall("AbstractText"):
                label = txt.get("Label", "").lower() or "full"
                content = "".join(txt.itertext())
                abstract_sections[label] = content
                abstract_full += content + " "
        rec["abstract"] = abstract_full.strip()
        rec["abstract_sections"] = abstract_sections

        authors = []
        for au in art.findall(".//Author"):
            last  = au.findtext("LastName") or ""
            fore  = au.findtext("ForeName") or au.findtext("Initials") or ""
            orcid = ""
            for idf in au.findall("Identifier"):
                if idf.get("Source") == "ORCID":
                    orcid = idf.text or ""
            affl = au.findtext(".//Affiliation") or ""
            authors.append({"last": last, "fore": fore, "orcid": orcid, "affiliation": affl})
        rec["authors"]      = authors
        rec["first_author"] = f'{authors[0]["last"]} {authors[0]["fore"]}' if authors else ""
        rec["all_authors"]  = "; ".join(f'{a["last"]} {a["fore"]}' for a in authors)

        journal_el = art.find("Journal")
        rec["journal_full"] = journal_el.findtext("Title") or "" if journal_el is not None else ""
        rec["journal"]      = journal_el.findtext("ISOAbbreviation") or rec["journal_full"] if journal_el is not None else ""
        jissue  = journal_el.find("JournalIssue") if journal_el is not None else None
        rec["volume"] = jissue.findtext("Volume") or "" if jissue is not None else ""
        rec["issue"]  = jissue.findtext("Issue") or "" if jissue is not None else ""
        pubdate = jissue.find("PubDate") if jissue is not None else None
        rec["year"] = (pubdate.findtext("Year") or pubdate.findtext("MedlineDate", "")[:4]
                       if pubdate is not None else "")
        rec["epub_date"] = ""
        for ah in art.findall("ArticleDate"):
            if ah.get("DateType") == "Electronic":
                rec["epub_date"] = f'{ah.findtext("Year","")}-{ah.findtext("Month","")}-{ah.findtext("Day","")}'
        rec["pages"] = art.findtext("Pagination/MedlinePgn") or ""

        rec["doi"] = ""
        rec["pmc"] = ""
        for aid in article.findall(".//ArticleId"):
            if aid.get("IdType") == "doi": rec["doi"] = aid.text or ""
            if aid.get("IdType") == "pmc": rec["pmc"] = aid.text or ""

        mesh_all, mesh_major = [], []
        for mh in article.findall(".//MeshHeading"):
            desc = mh.find("DescriptorName")
            if desc is not None:
                term = desc.text or ""
                mesh_all.append(term)
                if desc.get("MajorTopicYN") == "Y":
                    mesh_major.append(term)
        rec["mesh_str"]       = "; ".join(mesh_all)
        rec["mesh_major_str"] = "; ".join(mesh_major)

        kws = [k.text or "" for k in article.findall(".//Keyword")]
        rec["keywords_str"] = "; ".join(kws)

        pts = [p.text or "" for p in art.findall(".//PublicationType")]
        rec["pub_types_str"] = "; ".join(pts)

        chems = [n.findtext("NameOfSubstance") or "" for n in article.findall(".//Chemical")]
        rec["chemicals_str"] = "; ".join(chems)

        grants = []
        for g in article.findall(".//Grant"):
            gi = g.findtext("GrantID") or ""
            ga = g.findtext("Agency") or ""
            if gi or ga: grants.append(f"{gi} ({ga})")
        rec["grants_str"] = "; ".join(grants)

        rec["conflict_of_interest"] = article.findtext(".//CoiStatement") or ""
        affils = list({a["affiliation"] for a in authors if a["affiliation"]})
        rec["affiliations_str"] = " | ".join(affils)
        rec["country"]    = med.findtext(".//Country") or ""
        langs = [l.text or "" for l in art.findall(".//Language")]
        rec["languages"]  = "; ".join(langs)
        rec["url_pubmed"] = f"https://pubmed.ncbi.nlm.nih.gov/{rec['pmid']}/"
        rec["open_access_pdf"] = f"https://doi.org/{rec['doi']}" if rec["doi"] else ""

        rec["citation_count"]       = 0
        rec["influential_citations"] = 0
        rec["impact_factor_est"]    = 0.0
        rec["cited_by_count"]       = 0
        rec["venue"]   = rec["journal_full"]
        rec["source"]  = "PubMed"
        out.append(rec)
    return out


def records_to_df(records: list) -> pd.DataFrame:
    rows = []
    for r in records:
        rows.append({
            "PMID":                   r.get("pmid", ""),
            "Title":                  r.get("title", ""),
            "First Author":           r.get("first_author", ""),
            "All Authors":            r.get("all_authors", ""),
            "Year":                   r.get("year", ""),
            "Journal (Full)":         r.get("journal_full", ""),
            "Journal (Abbrev)":       r.get("journal", ""),
            "Volume":                 r.get("volume", ""),
            "Issue":                  r.get("issue", ""),
            "Pages":                  r.get("pages", ""),
            "DOI":                    r.get("doi", ""),
            "PMC":                    r.get("pmc", ""),
            "Abstract (Full)":        r.get("abstract", ""),
            "MeSH Major":             r.get("mesh_major_str", ""),
            "MeSH All":               r.get("mesh_str", ""),
            "Keywords":               r.get("keywords_str", ""),
            "Publication Types":      r.get("pub_types_str", ""),
            "Chemicals":              r.get("chemicals_str", ""),
            "Grants":                 r.get("grants_str", ""),
            "COI":                    r.get("conflict_of_interest", ""),
            "Affiliations":           r.get("affiliations_str", ""),
            "Country":                r.get("country", ""),
            "Languages":              r.get("languages", ""),
            "Citations":              r.get("citation_count", 0),
            "Influential Citations":  r.get("influential_citations", 0),
            "Impact Factor Est":      r.get("impact_factor_est", 0.0),
            "Venue":                  r.get("venue", ""),
            "PubMed URL":             r.get("url_pubmed", ""),
            "Open Access PDF":        r.get("open_access_pdf", ""),
        })
    return pd.DataFrame(rows)


# ── Column group definitions for MSL Excel ───────────────────────
_COL_GROUPS = {
    "IDENTIFICATION": {
        "color": "1A3A5C",
        "cols": ["PMID", "Study Title", "First Author", "All Authors", "Year",
                 "Journal (Full)", "Journal (Abbrev)", "Volume", "Issue",
                 "Pages", "DOI", "PMC", "Trial Registration",
                 "PubMed URL", "Open Access PDF"],
    },
    "STUDY DESIGN": {
        "color": "1B5E3B",
        "cols": ["Study Design", "Study Phase", "Study Type",
                 "Sample Size (N)", "Follow-up Duration",
                 "Randomisation", "Blinding", "Control Type",
                 "ITT / Per-Protocol", "Countries / Sites"],
    },
    "POPULATION & INTERVENTION": {
        "color": "5C3A1A",
        "cols": ["Population / Indication", "Inclusion Criteria",
                 "Exclusion Criteria", "Intervention", "Dose / Regimen",
                 "Comparator / Control", "Background Therapy"],
    },
    "ENDPOINTS & OUTCOMES": {
        "color": "3A1A5C",
        "cols": ["Primary Endpoint", "Secondary Endpoints",
                 "Exploratory Endpoints", "Key Efficacy Outcomes",
                 "Effect Size (Primary)", "95% CI", "P-value",
                 "Relative Risk Reduction", "Absolute Risk Reduction",
                 "NNT", "NNH", "Statistical Method", "Subgroup Analyses"],
    },
    "SAFETY": {
        "color": "5C1A1A",
        "cols": ["Safety Population (N)", "Any AE (%)",
                 "Serious AE (%)", "Discontinuation due to AE (%)",
                 "Key AEs of Interest", "Deaths (%)"],
    },
    "CONTEXT & QUALITY": {
        "color": "1A4A5C",
        "cols": ["Limitations", "Funding Source", "COI",
                 "Guideline Relevance", "MSL Key Message"],
    },
    "BIBLIOMETRICS": {
        "color": "2D2D2D",
        "cols": ["Citations", "Influential Citations", "Impact Factor Est",
                 "Publication Types", "MeSH Major", "MeSH All", "Keywords",
                 "Chemicals", "Grants", "Abstract (Full)"],
    },
}


def _get_ordered_columns() -> list:
    cols = []
    for grp in _COL_GROUPS.values():
        for c in grp["cols"]:
            if c not in cols:
                cols.append(c)
    return cols


def _col_group_for(col: str) -> tuple[str, str]:
    for grp_name, grp_data in _COL_GROUPS.items():
        if col in grp_data["cols"]:
            return grp_name, grp_data["color"]
    return "OTHER", "444444"


def to_excel_bytes(records: list, extracted_rows: list = None,
                   include_abstract: bool = True) -> bytes:
    """
    Build a professional MSL-grade Excel workbook.
    - Sheet 1: Full Evidence Table (raw PubMed + LLM extracted, colour-coded by column group)
    - Sheet 2: Raw PubMed Data
    - Sheet 3: Summary Stats
    """
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()

    # ── Sheet 1: Full MSL Evidence Table ─────────────────────────
    ws1 = wb.active
    ws1.title = "MSL Evidence Table"

    # Merge PubMed records with LLM extracted rows
    merged_rows = _merge_records(records, extracted_rows)
    ordered_cols = _get_ordered_columns()
    if not include_abstract and "Abstract (Full)" in ordered_cols:
        ordered_cols.remove("Abstract (Full)")

    # Filter to columns that have any data
    present_cols = [c for c in ordered_cols
                    if any(str(row.get(c, "")).strip() for row in merged_rows)]

    _write_evidence_sheet(ws1, merged_rows, present_cols)

    # ── Sheet 2: Raw PubMed ───────────────────────────────────────
    ws2 = wb.create_sheet("Raw PubMed Data")
    raw_df = records_to_df(records)
    _write_raw_sheet(ws2, raw_df)

    # ── Sheet 3: Summary Stats ────────────────────────────────────
    ws3 = wb.create_sheet("Summary Statistics")
    _write_summary_sheet(ws3, records)

    wb.save(buf)
    return buf.getvalue()


def _merge_records(records: list, extracted_rows: list = None) -> list:
    """Merge PubMed raw fields with LLM-extracted fields by PMID."""
    # Build lookup of extracted rows by PMID
    ext_by_pmid = {}
    if extracted_rows:
        for row in extracted_rows:
            pmid = str(row.get("PMID", ""))
            if pmid:
                ext_by_pmid[pmid] = row

    merged = []
    for rec in records:
        pmid = str(rec.get("pmid", ""))
        row  = {}

        # Raw PubMed fields
        row["PMID"]                = rec.get("pmid", "")
        row["Title"]               = rec.get("title", "")
        row["First Author"]        = rec.get("first_author", "")
        row["All Authors"]         = rec.get("all_authors", "")
        row["Year"]                = rec.get("year", "")
        row["Journal (Full)"]      = rec.get("journal_full", "")
        row["Journal (Abbrev)"]    = rec.get("journal", "")
        row["Volume"]              = rec.get("volume", "")
        row["Issue"]               = rec.get("issue", "")
        row["Pages"]               = rec.get("pages", "")
        row["DOI"]                 = rec.get("doi", "")
        row["PMC"]                 = rec.get("pmc", "")
        row["PubMed URL"]          = rec.get("url_pubmed", "")
        row["Open Access PDF"]     = rec.get("open_access_pdf", "")
        row["Abstract (Full)"]     = rec.get("abstract", "")
        row["MeSH Major"]          = rec.get("mesh_major_str", "")
        row["MeSH All"]            = rec.get("mesh_str", "")
        row["Keywords"]            = rec.get("keywords_str", "")
        row["Publication Types"]   = rec.get("pub_types_str", "")
        row["Chemicals"]           = rec.get("chemicals_str", "")
        row["Grants"]              = rec.get("grants_str", "")
        row["Conflict of Interest"]= rec.get("conflict_of_interest", "")
        row["Country"]             = rec.get("country", "")
        row["Citations"]           = rec.get("citation_count", 0)
        row["Influential Citations"]= rec.get("influential_citations", 0)
        row["Impact Factor Est"]   = rec.get("impact_factor_est", 0.0)

        # LLM-extracted fields (if available for this PMID)
        ext = ext_by_pmid.get(pmid, {})
        def g(k, *aliases):
            """Get from extracted with multiple key aliases."""
            for key in [k] + list(aliases):
                v = ext.get(key, "")
                if v and v != "NR": return v
            return ext.get(k, "")

        row["Study Design"]                  = g("Study Design")
        row["Study Phase"]                   = g("Study Phase")
        row["Study Type"]                    = g("Study Type", "Publication Types")
        row["Sample Size (N)"]               = g("Sample Size", "Sample Size (N)")
        row["Follow-up Duration"]            = g("Follow-up Duration", "Follow-up")
        row["Randomisation"]                 = g("Randomisation")
        row["Blinding"]                      = g("Blinding")
        row["Control Type"]                  = g("Control Type")
        row["Population / Indication"]       = g("Population / Indication", "Population")
        row["Inclusion Criteria"]            = g("Inclusion Criteria")
        row["Exclusion Criteria"]            = g("Exclusion Criteria")
        row["Intervention"]                  = g("Intervention")
        row["Comparator / Control"]          = g("Comparator / Control", "Comparator")
        row["Background Therapy"]            = g("Background Therapy")
        row["Dose / Regimen"]                = g("Dose / Regimen")
        row["Countries / Sites"]             = g("Countries / Sites") or rec.get("country","")
        row["Trial Registration"]            = g("Trial Registration")
        row["Primary Endpoint"]              = g("Primary Endpoint")
        row["Secondary Endpoints"]           = g("Secondary Endpoints")
        row["Exploratory Endpoints"]         = g("Exploratory Endpoints")
        row["Key Efficacy Outcomes"]         = g("Key Efficacy Outcomes", "Key Outcomes")
        row["Effect Size (Primary)"]         = g("Effect Size (Primary)", "Effect Size")
        row["95% CI"]                        = g("95% CI")
        row["P-value"]                       = g("P-value")
        row["NNT"]                           = g("NNT")
        row["NNH"]                           = g("NNH")
        row["Relative Risk Reduction"]       = g("Relative Risk Reduction")
        row["Absolute Risk Reduction"]       = g("Absolute Risk Reduction")
        row["Statistical Method"]            = g("Statistical Method")
        row["ITT / Per-Protocol"]            = g("ITT / Per-Protocol")
        row["Subgroup Analyses"]             = g("Subgroup Analyses")
        row["Safety Population (N)"]         = g("Safety Population (N)", "Safety Population")
        row["Any AE (%)"]                    = g("Any AE (%)")
        row["Serious AE (%)"]                = g("Serious AE (%)")
        row["Discontinuation due to AE (%)"] = g("Discontinuation due to AE (%)", "Discontinuation AE (%)")
        row["Key AEs of Interest"]           = g("Key AEs of Interest", "Safety")
        row["Deaths (%)"]                    = g("Deaths (%)", "Deaths")
        row["Limitations"]                   = g("Limitations")
        row["Funding Source"]                = g("Funding Source", "Funding")
        row["Guideline Relevance"]           = g("Guideline Relevance")
        row["MSL Key Message"]               = g("MSL Key Message", "MSL Notes")

        merged.append(row)
    return merged


def _write_evidence_sheet(ws, rows: list, cols: list):
    """Write the full MSL evidence table with colour-coded column groups."""
    thin = Side(style="thin", color="334155")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: Group header ───────────────────────────────────────
    current_grp = None
    grp_start   = 1
    group_spans = []  # (col_start, col_end, grp_name, color)

    for ci, col in enumerate(cols, 1):
        grp_name, grp_color = _col_group_for(col)
        if grp_name != current_grp:
            if current_grp is not None:
                group_spans.append((grp_start, ci - 1, current_grp,
                                    _COL_GROUPS.get(current_grp, {}).get("color", "333333")))
            current_grp = grp_name
            grp_start   = ci
    if current_grp:
        group_spans.append((grp_start, len(cols), current_grp,
                            _COL_GROUPS.get(current_grp, {}).get("color", "333333")))

    # Write group header row (row 1)
    for (cs, ce, gname, gcolor) in group_spans:
        cell = ws.cell(row=1, column=cs, value=gname)
        cell.fill = PatternFill("solid", fgColor=gcolor)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        if cs != ce:
            ws.merge_cells(start_row=1, start_column=cs,
                           end_row=1, end_column=ce)

    # ── Row 2: Column header ──────────────────────────────────────
    for ci, col in enumerate(cols, 1):
        _, grp_color = _col_group_for(col)
        # Slightly lighter shade for column header
        cell = ws.cell(row=2, column=ci, value=col)
        cell.fill = PatternFill("solid", fgColor=grp_color)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 36

    # ── Data rows ─────────────────────────────────────────────────
    for ri, row in enumerate(rows, 3):
        fill_color = "1E293B" if ri % 2 == 0 else "0F172A"
        for ci, col in enumerate(cols, 1):
            val  = row.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.font      = Font(color="CBD5E1", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = border

            # Hyperlink for URL columns
            if col in ("PubMed URL", "Open Access PDF", "DOI") and val:
                href = val if val.startswith("http") else f"https://doi.org/{val}"
                cell.hyperlink = href
                cell.font = Font(color="38BDF8", size=9, underline="single")

            # Highlight non-NR extracted fields
            if val and val not in ("NR", "Not Reported", "Not Available", ""):
                _, grp_color = _col_group_for(col)
                if col in ["Primary Endpoint", "Effect Size (Primary)",
                           "P-value", "Key Efficacy Outcomes", "Comparator / Control"]:
                    cell.font = Font(color="F0FDF4", size=9, bold=True)

    # ── Column widths ─────────────────────────────────────────────
    wide_cols  = {"Title", "Abstract (Full)", "Key Efficacy Outcomes",
                  "Primary Endpoint", "Secondary Endpoints", "Limitations",
                  "Population / Indication", "Key AEs of Interest",
                  "Subgroup Analyses", "All Authors"}
    narrow_cols = {"PMID", "Year", "Volume", "Issue", "Pages",
                   "Citations", "P-value", "Study Phase"}

    for ci, col in enumerate(cols, 1):
        letter = get_column_letter(ci)
        if col in wide_cols:
            ws.column_dimensions[letter].width = 45
        elif col in narrow_cols:
            ws.column_dimensions[letter].width = 10
        else:
            ws.column_dimensions[letter].width = 22

    ws.freeze_panes = "C3"
    ws.sheet_view.showGridLines = False


def _write_raw_sheet(ws, df: pd.DataFrame):
    """Write raw PubMed data with simple clean styling."""
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    thin = Side(style="thin", color="334155")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 30

    for ri, row_data in enumerate(df.itertuples(index=False), 2):
        fill_color = "1E293B" if ri % 2 == 0 else "0F172A"
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.font      = Font(color="CBD5E1", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = border

    for ci, col in enumerate(df.columns, 1):
        letter = get_column_letter(ci)
        ws.column_dimensions[letter].width = 40 if col in (
            "Title", "Abstract (Full)", "MeSH All") else 18

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _write_summary_sheet(ws, records: list):
    """Write a summary statistics sheet."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    h_fill = PatternFill("solid", fgColor="1A3A5C")
    h_font = Font(bold=True, color="FFFFFF", size=11)
    v_font = Font(color="E2E8F0", size=10)
    s_font = Font(bold=True, color="38BDF8", size=13)
    bg     = PatternFill("solid", fgColor="0F172A")

    def write_row(r, label, value, is_header=False):
        ca = ws.cell(row=r, column=1, value=label)
        cb = ws.cell(row=r, column=2, value=value)
        for c in (ca, cb):
            c.fill = h_fill if is_header else bg
            c.font = h_font if is_header else (s_font if isinstance(value, (int, float)) else v_font)
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 22

    write_row(1, "MSL EVIDENCE SUMMARY", "", is_header=True)
    write_row(2, "Total Records", len(records))

    years = [int(r.get("year", 0) or 0) for r in records if r.get("year")]
    if years:
        write_row(3, "Year Range", f"{min(years)} – {max(years)}")
        from collections import Counter
        top_years = Counter(years).most_common(3)
        write_row(4, "Most Active Years", ", ".join(f"{y}({c})" for y,c in top_years))

    journals = [r.get("journal_full","") for r in records if r.get("journal_full")]
    if journals:
        from collections import Counter
        top_j = Counter(journals).most_common(5)
        write_row(5, "Top Journals (by volume)", "")
        for i, (j, c) in enumerate(top_j, 6):
            write_row(i, f"  {j}", c)

    cits = [r.get("citation_count", 0) for r in records]
    if any(cits):
        write_row(11, "Total Citations", sum(cits))
        write_row(12, "Median Citations", sorted(cits)[len(cits)//2])
        write_row(13, "Max Citations", max(cits))

    countries = [r.get("country","") for r in records if r.get("country")]
    if countries:
        from collections import Counter
        top_c = Counter(countries).most_common(5)
        write_row(14, "Top Countries", ", ".join(f"{c}({n})" for c,n in top_c))

    pub_types_all = []
    for r in records:
        for pt in r.get("pub_types_str","").split(";"):
            pt = pt.strip()
            if pt: pub_types_all.append(pt)
    if pub_types_all:
        from collections import Counter
        top_pt = Counter(pub_types_all).most_common(5)
        write_row(15, "Top Publication Types", "")
        for i, (pt, c) in enumerate(top_pt, 16):
            write_row(i, f"  {pt}", c)


# ── Clinical Trials Excel export ──────────────────────────────────
_CT_COL_GROUPS = {
    "IDENTIFICATION": {
        "color": "1A3A5C",
        "cols": ["nct_id", "title", "official_title", "sponsor", "status", "phase", "study_type"],
    },
    "TIMELINE & SIZE": {
        "color": "1B5E3B",
        "cols": ["start_date", "completion_date", "enrollment", "countries"],
    },
    "CLINICAL": {
        "color": "5C3A1A",
        "cols": ["conditions", "interventions", "primary_outcome"],
    },
    "SUMMARY": {
        "color": "3A1A5C",
        "cols": ["summary"],
    },
    "LINK": {
        "color": "2D2D2D",
        "cols": ["url"],
    },
}

_CT_COL_LABELS = {
    "nct_id":           "NCT ID",
    "title":            "Brief Title",
    "official_title":   "Official Title",
    "sponsor":          "Lead Sponsor",
    "status":           "Overall Status",
    "phase":            "Phase",
    "study_type":       "Study Type",
    "start_date":       "Start Date",
    "completion_date":  "Primary Completion",
    "enrollment":       "Enrollment (N)",
    "countries":        "Countries",
    "conditions":       "Conditions",
    "interventions":    "Interventions",
    "primary_outcome":  "Primary Outcome Measure",
    "summary":          "Brief Summary",
    "url":              "ClinicalTrials.gov URL",
}

_CT_COL_WIDTHS = {
    "nct_id": 16, "title": 40, "official_title": 45,
    "sponsor": 28, "status": 20, "phase": 14, "study_type": 18,
    "start_date": 14, "completion_date": 18, "enrollment": 14,
    "countries": 22, "conditions": 30, "interventions": 35,
    "primary_outcome": 40, "summary": 50, "url": 30,
}

_STATUS_COLORS = {
    "RECRUITING":            "0D6E3A",
    "COMPLETED":             "1A3A5C",
    "ACTIVE_NOT_RECRUITING": "7A5200",
    "TERMINATED":            "7A1A1A",
    "NOT_YET_RECRUITING":    "3D3D3D",
    "WITHDRAWN":             "4A4A4A",
    "SUSPENDED":             "6B3A00",
}


def ct_to_excel_bytes(studies: list) -> bytes:
    """
    Build a professional MSL-grade Excel workbook for ClinicalTrials.gov data.
    - Sheet 1: Full styled trials table with colour-coded groups
    - Sheet 2: Summary statistics (phase breakdown, status counts, top sponsors)
    """
    if not studies:
        raise ValueError("No studies to export")

    buf = io.BytesIO()
    wb  = openpyxl.Workbook()

    # ── Sheet 1: Clinical Trials Table ───────────────────────────
    ws1 = wb.active
    ws1.title = "Clinical Trials"
    _write_ct_table(ws1, studies)

    # ── Sheet 2: Summary ─────────────────────────────────────────
    ws2 = wb.create_sheet("Summary Statistics")
    _write_ct_summary(ws2, studies)

    wb.save(buf)
    return buf.getvalue()


def _write_ct_table(ws, studies: list):
    """Write the styled clinical trials table."""
    from openpyxl.utils import get_column_letter

    thin   = Side(style="thin", color="2D3748")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Build ordered column list
    all_cols = []
    col_group_map = {}
    for grp_name, grp_data in _CT_COL_GROUPS.items():
        for c in grp_data["cols"]:
            all_cols.append(c)
            col_group_map[c] = (grp_name, grp_data["color"])

    # ── Row 1: Group header ───────────────────────────────────────
    current_grp = None
    grp_start   = 1
    spans = []
    for ci, col in enumerate(all_cols, 1):
        grp_name, grp_color = col_group_map[col]
        if grp_name != current_grp:
            if current_grp is not None:
                spans.append((grp_start, ci - 1, current_grp,
                               col_group_map[all_cols[grp_start-1]][1]))
            current_grp = grp_name
            grp_start   = ci
    if current_grp:
        spans.append((grp_start, len(all_cols), current_grp,
                       col_group_map[all_cols[grp_start-1]][1]))

    for cs, ce, gname, gcolor in spans:
        cell = ws.cell(row=1, column=cs, value=gname)
        cell.fill      = PatternFill("solid", fgColor=gcolor)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cs != ce:
            ws.merge_cells(start_row=1, start_column=cs, end_row=1, end_column=ce)

    # ── Row 2: Column labels ──────────────────────────────────────
    for ci, col in enumerate(all_cols, 1):
        _, grp_color = col_group_map[col]
        label = _CT_COL_LABELS.get(col, col.replace("_", " ").title())
        cell  = ws.cell(row=2, column=ci, value=label)
        cell.fill      = PatternFill("solid", fgColor=grp_color)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 36

    # ── Data rows ─────────────────────────────────────────────────
    for ri, study in enumerate(studies, 3):
        fill_color = "1E293B" if ri % 2 == 0 else "0F172A"
        status_raw = str(study.get("status", "")).upper().replace(" ", "_")

        for ci, col in enumerate(all_cols, 1):
            val = study.get(col, "")
            # Sanitise: convert any non-string to string, strip None
            if val is None:
                val = ""
            elif isinstance(val, (list, dict)):
                val = str(val)
            else:
                val = str(val)

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Status column gets colour-coded background
            if col == "status":
                sc = _STATUS_COLORS.get(status_raw, "2D3748")
                cell.fill = PatternFill("solid", fgColor=sc)
                cell.font = Font(color="FFFFFF", bold=True, size=9)
            elif col == "nct_id" and val:
                cell.fill      = PatternFill("solid", fgColor=fill_color)
                cell.font      = Font(color="38BDF8", size=9, bold=True)
                cell.hyperlink = study.get("url", f"https://clinicaltrials.gov/study/{val}")
            elif col == "url" and val:
                cell.fill      = PatternFill("solid", fgColor=fill_color)
                cell.font      = Font(color="38BDF8", size=9, underline="single")
                cell.hyperlink = val
            elif col == "phase":
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(color="F59E0B", size=9, bold=True)
            elif col == "enrollment":
                cell.fill      = PatternFill("solid", fgColor=fill_color)
                cell.font      = Font(color="10B981", size=9, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(color="CBD5E1", size=9)

        ws.row_dimensions[ri].height = 55

    # ── Column widths ─────────────────────────────────────────────
    for ci, col in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = _CT_COL_WIDTHS.get(col, 20)

    ws.freeze_panes = "C3"
    ws.sheet_view.showGridLines = False


def _write_ct_summary(ws, studies: list):
    """Write summary statistics for the CT results."""
    from collections import Counter
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18

    bg     = PatternFill("solid", fgColor="0F172A")
    h_fill = PatternFill("solid", fgColor="1A3A5C")
    h_font = Font(bold=True, color="FFFFFF", size=11)
    v_font = Font(color="CBD5E1", size=10)
    n_font = Font(color="38BDF8", size=12, bold=True)

    def hdr(r, text):
        c = ws.cell(row=r, column=1, value=text)
        c.fill = h_fill
        c.font = h_font
        c.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 24

    def row(r, label, val, pct=None):
        ca = ws.cell(row=r, column=1, value=label)
        cb = ws.cell(row=r, column=2, value=val)
        ca.fill = cb.fill = bg
        ca.font = v_font
        cb.font = n_font
        ca.alignment = Alignment(vertical="center")
        cb.alignment = Alignment(vertical="center")
        if pct is not None:
            cc = ws.cell(row=r, column=3, value=f"{pct:.1f}%")
            cc.fill = bg
            cc.font = Font(color="64748B", size=9)
        ws.row_dimensions[r].height = 22

    r = 1
    hdr(r, "CLINICAL TRIALS SUMMARY"); r += 1
    row(r, "Total Studies Found", len(studies)); r += 1

    # Status breakdown
    r += 1
    hdr(r, "BY STATUS"); r += 1
    statuses = Counter(s.get("status", "Unknown") for s in studies)
    for status, count in statuses.most_common():
        row(r, f"  {status}", count, count / len(studies) * 100)
        r += 1

    # Phase breakdown
    r += 1
    hdr(r, "BY PHASE"); r += 1
    phases = Counter(s.get("phase", "Not Specified") or "Not Specified" for s in studies)
    for phase, count in phases.most_common():
        row(r, f"  {phase}", count, count / len(studies) * 100)
        r += 1

    # Top sponsors
    r += 1
    hdr(r, "TOP SPONSORS"); r += 1
    sponsors = Counter(s.get("sponsor", "") for s in studies if s.get("sponsor"))
    for sponsor, count in sponsors.most_common(8):
        row(r, f"  {sponsor[:40]}", count)
        r += 1

    # Enrollment stats
    enrollments = []
    for s in studies:
        try:
            e = int(s.get("enrollment") or 0)
            if e > 0:
                enrollments.append(e)
        except (ValueError, TypeError):
            pass
    if enrollments:
        r += 1
        hdr(r, "ENROLLMENT STATISTICS"); r += 1
        row(r, "  Total Participants",    sum(enrollments)); r += 1
        row(r, "  Median Enrollment",     sorted(enrollments)[len(enrollments)//2]); r += 1
        row(r, "  Largest Trial",         max(enrollments)); r += 1
        row(r, "  Trials with N > 1000",  sum(1 for e in enrollments if e > 1000)); r += 1
